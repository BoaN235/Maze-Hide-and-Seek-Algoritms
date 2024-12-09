import pygame
from CellClass import Cell
from ActorClass import Actor, CauseOfDeath
from PredActor import PredActor
from PreyActor import PreyActor
from inputs import Button, InputBox, Text, Slider, TransparentBox
import ast
import json
from random import seed
import random
from openpyxl import Workbook , load_workbook
import time
from Notification import Email
import os
import zlib

class SimState:
    def __init__(self):
        self.MaxActions = 100
        self.WIDTH, self.HEIGHT = 1202, 902
        self.TILE = 25
        self.cols, self.rows = self.WIDTH // self.TILE, self.HEIGHT // self.TILE        
        self.generation = 0
        self.pred_score = 0
        self.generation_actors = []
        self.killed_actors = []
        self.setting = False
        self.speed = 100
        self.max_preds = 20
        self.max_preys = 100  
        self.preds = 20
        self.preys = 100     
        self.Actors = []
        self.max_generations = 1001
        self.running = True
        self.path = "data/simdata.xlsx"
        self.workbook = self.create_workbook()
        self.current_row = 2
        self.sheet = self.create_sheet()
        self.start_time = time.time()
        self.end_time = time.time()
        self.time = time.time()
        self.time_in_ms = self.time * 1000
        self.save_every = 10
        self.food_list = []
        self.action_step = 0
        self.state = []


    def load_walls(self):
        try:
            with open("lastmaze.txt", "r") as file:
                file_contents = file.readlines()
                self.TILE = int(file_contents[0].strip())
                if len(file_contents) > 1:
                    # Process the remaining lines
                    walls_data = [ast.literal_eval(line.strip()) for line in file_contents[1:]]
        except (FileNotFoundError, ValueError, SyntaxError) as e:
            print(f"Error loading maze state: {e}")
            self.TILE = 100
            walls_data = []
        
        return walls_data

    def save_walls(self):
        pass

    def create_grid(self):
            # Initialize grid cells
        self.grid_cells = [Cell(col, row, self) for row in range(self.rows) for col in range(self.cols)]
        # Load saved walls if available
        
        walls_data = self.load_walls()
        
        if walls_data:
            for cell, walls in zip(self.grid_cells, walls_data):
                cell.save_gen(walls)
    def create_actors(self):
        total_actors = self.preds + self.preys
        spawn_indices = random.sample(range(len(self.grid_cells)), total_actors)

        for i in range(self.preds):
            spawn_index = spawn_indices.pop()
            self.Actors.append(PredActor(self, spawn_index, i))

        for i in range(self.preys):
            spawn_index = spawn_indices.pop()
            self.Actors.append(PreyActor(self, spawn_index, i))


        self.generation_actors = self.Actors


    def load(self):
        # Load the saved state
        with open("state.json", 'r') as f:
                state = json.load(f)
            
        self.MaxActions = state['max_actions']
        self.WIDTH = state['width']
        self.HEIGHT = state['height']
        self.TILE = state['tile']
        self.cols = state['cols']
        self.rows = state['rows']
        self.generation = state['generation']
        
        self.grid_cells = [Cell.from_dict(cell_data) for cell_data in state['grid_cells']]
        self.Actors = [Actor.from_dict(actor_data) for actor_data in state['actors']]

        pass

    def save(self):
        # Save the current state
        cells = [cell.to_list() for cell in self.grid_cells]
        Actors = [actor.to_list() for actor in self.Actors]
        states = [
            self.MaxActions,
            self.WIDTH,
            self.HEIGHT,
            self.TILE,
            self.cols,
            self.rows,
            self.generation,
            cells,
            Actors,
            [cell.food for cell in self.grid_cells]
        ]
        self.state.append(states)





    def save_gen_stats(self):
        # Save the current state
        
        save = []

        for x in self.Actors:
            if x.dead:
                self.pred_score += 1
        for a in self.Actors:
            save.append(a.actions)
            
        directory = "data/Replays"
        filename = f"{directory}/replay_{self.generation}.json"

        # Ensure the directory exists
        os.makedirs(directory, exist_ok=True)

        try:
            with open(filename, 'w') as file:
                json.dump([save], file, indent=4)
            print(f"Saved simulation state to {filename}")
        except FileNotFoundError:
            print(f"File not found: {filename}")

        preds = 0
        preys = 0
        starved_prey = 0
        starved_pred = 0

        for a in self.Actors:
            if isinstance(a, PredActor):
                if a.cause_of_death == CauseOfDeath.STARVATION:
                    starved_pred += 1
                if not a.dead:
                    preds += 1
            elif isinstance(a, PreyActor):
                if a.cause_of_death == CauseOfDeath.STARVATION:
                    starved_prey += 1
                if not a.dead:
                    preys += 1


        self.sheet["A" + str(self.current_row)] = self.generation
        self.sheet["B" + str(self.current_row)] = "wins_data"
        self.sheet["C" + str(self.current_row)] = preys
        self.sheet["D" + str(self.current_row)] = preds 
        self.sheet["E" + str(self.current_row)] = self.time_in_ms # yes
        self.sheet["F" + str(self.current_row)] = starved_prey #prey starved
        self.sheet["G" + str(self.current_row)] = starved_pred #predator starved

        self.current_row += 1


        self.workbook.save(self.path)

    def replay(self):
        current_replay = 1

        pass

    def create_workbook(self):
        workbook = Workbook()
        
        return workbook
    
    def create_sheet(self):
        sheet = self.workbook.active
        sheet.freeze_panes = "A1"
        headers = ["Generation", "Winner", "Prey", "Pred", "Time(ms)","prey starved","predator starved"]
        sheet["A1"] = headers[0]
        sheet["B1"] = headers[1]
        sheet["C1"] = headers[2]
        sheet["D1"] = headers[3]
        sheet["E1"] = headers[4]
        sheet["F1"] = headers[5]
        sheet["G1"] = headers[6]
        sheet.title = "Data"
        self.workbook.save(self.path)
        return sheet

    def end_sim(self):
        email = Email()
        email.send_email()
        self.running = False
        self.start_review()
        with open("state.json", 'wb') as f:
            f.write(str(self.state))


    

    def reset_generation(self):
        # Reset the generation
        for i, cell in enumerate(self.grid_cells):
            cell.food = self.food_list[i]

        
        self.action_step = 0
        self.end_time = time.time()
        self.time = self.end_time - self.start_time
        
                # Convert time to milliseconds
        self.time_in_ms = self.time * 1000
        # Convert time to seconds
        time_in_s = self.time
        self.save()
        self.save_gen_stats()


        self.generation += 1
        print(f"Generation: {self.generation}  time: {time_in_s:.2f} seconds ({self.time_in_ms:.0f} milliseconds)")
        self.pred_score = 0

        if self.generation >= self.max_generations:
            self.end_sim()
        for x in self.Actors: # Respawn all dead actors
            x.reset()
        self.start_time = time.time()
        self.preds = self.max_preds
        self.preys = self.max_preys   
        pass
    def settings(self):
        self.setting = not self.setting

    def draw_settings_screen(self, screen, settings_bar, font, slider, input_box, view_button):
        if self.setting:
            settings_bar.draw_box()
            slider.draw_slider()
            input_box.draw_input_box()
            view_button.draw_button()

    def generate_sim(self):
        # Game loop
        seed(132)
        RES = self.WIDTH, self.HEIGHT
        FONT_SIZE = 32

        # Pygame setup
        pygame.init()
        sc = pygame.display.set_mode(RES)
        clock = pygame.time.Clock()
        font = pygame.font.Font(None, 32)

        sc.fill(pygame.Color(50, 50, 50))
        


        # Initialize the grid
        self.grid_cells = [Cell(col, row, self) for row in range(self.rows) for col in range(self.cols)]
        current_cell =  self.grid_cells[0]
        stack = []

        done = False

        # Main loop
        while self.running:

            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    exit()

            # Draw all cells
            [cell.draw(sc) for cell in self.grid_cells]

            # Mark the current cell as visited and draw it
            current_cell.visited = True
            current_cell.draw_current_cell(sc)

            # Check for the next cell to move to
            next_cell = current_cell.check_neighbors(self.grid_cells)
            if next_cell:
                next_cell.visited = True
                stack.append(current_cell)
                Cell.remove_walls(current_cell, next_cell)
                current_cell = next_cell
            elif stack:
                current_cell = stack.pop()

            # Check if all cells are visited and if so, write to the file
            if all(cell.visited for cell in self.grid_cells):
                if not done:
                    print("All cells visited, writing to file...")
                    with open("lastmaze.txt", "w") as file:
                        file.write(str(self.TILE) + '\n')
                        for x in self.grid_cells:
                            file.write(str(x.walls) + '\n')
                    print("File written successfully.")
                    done = True
                    pygame.quit()
                    exit()

            # Update display
            pygame.display.flip()
            clock.tick(1000)  # Adjust the tick rate for smoother gameplay

    def view(self):
        if self.grid_cells[0].current_view == 'food':
            for cell in self.grid_cells:
                cell.current_view = None
        else:
            for cell in self.grid_cells:
                cell.current_view = 'food'
        print('View button clicked')


    def start_sim(self):
        self.create_grid()
        self.create_actors()
        # Game loop
        RES = self.WIDTH, self.HEIGHT
        FONT_SIZE = 24
        
        for cell in self.grid_cells:
            food = random.randint(1, 10)
            self.food_list.append(food)
            cell.food = food

        
        pygame.init()
        sc = pygame.display.set_mode(RES)
        font = pygame.font.Font(None, FONT_SIZE)
        clock = pygame.time.Clock()
        
        # Create settings bar
        settings_bar_rect = pygame.Rect(0, 0, 200, self.HEIGHT)
        settings_bar = TransparentBox(sc, settings_bar_rect, (0, 0, 0), 128)

        # Create settings button
        settings_button_rect = pygame.Rect(10, 10, 180, 40)
        settings_button = Button(sc, (0, 0, 255), "Settings", settings_button_rect, font, self.settings)

        view_button_rect = pygame.Rect(10, 160, 180, 40)
        view_button = Button(sc, (0, 0, 255), "View", view_button_rect, font, self.view)


        # Create slider for speed control
        slider = Slider(sc, (10, 60), 180, 20, 1, 2000, self.speed)

        # Create input box for precise speed control
        input_data = {
            'color_active': (0, 255, 0),
            'color_passive': (255, 0, 0),
            'text': str(self.speed),
            'active': False
        }
        input_box_rect = pygame.Rect(10, 100, 180, 32)
        input_box = InputBox(sc, input_data, font, input_box_rect)

        while self.running:
            #sc.fill((30, 30, 30))
            sc.fill((0, 0, 0)) #cool mode
            #sc.fill((255, 255, 255)) #try it I DARE YOU



            if self.action_step >= self.MaxActions:
                self.reset_generation()
                self.Actors = self.generation_actors
            for a in self.Actors:
                a.step()
            self.action_step += 1


            # Draw all cells
            [cell.draw(sc) for cell in self.grid_cells]
            
            
            settings_button.draw_button()

            # Draw the settings bar and its elements
            self.draw_settings_screen(sc, settings_bar, font, slider, input_box, view_button)

            
            
            # Handle events
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    exit()
                settings_button.handle_event(event)
                view_button.handle_event(event)
                self.speed = int(input_box.handle_event(event))


            pygame.display.flip()
            clock.tick(self.speed)  # Adjust the tick rate for smoother gameplay    
    
    def start_review(self):
        # Game loop
        RES = self.WIDTH, self.HEIGHT
        FONT_SIZE = 34
        


        pygame.init()
        sc = pygame.display.set_mode(RES)
        font = pygame.font.Font(None, FONT_SIZE)
        titlefont = pygame.font.Font(None, 68)
        clock = pygame.time.Clock()
        
        texts = [Text(sc,  (self.WIDTH // 2, self.HEIGHT // 2 - 100), "Finished", titlefont, (255, 255, 255)), Text(sc,  (self.WIDTH // 2, self.HEIGHT // 2), "Please look at the exel spreadsheet", font, (255, 255, 255)), Text(sc,  (self.WIDTH // 2, self.HEIGHT // 2 + 100), "Please do not close this tab", font, (255, 0, 0))]
        buttons = [Button(sc, (0, 0, 255), "Next", pygame.Rect(self.WIDTH // 2 - 100, self.HEIGHT // 2 + 200, 200, 50), font, self.replay)]#, Button(sc, (0, 0, 255), "Back", pygame.Rect(self.WIDTH // 2 - 100, self.HEIGHT // 2 + 300, 200, 50), self.replay)
        while True:
            #sc.fill((30, 30, 30))
            sc.fill((0, 0, 0)) #cool mode
            #sc.fill((255, 255, 255)) try it I DARE YOU
            for x in texts:
                x.draw_text()
            for x in buttons:
                x.draw_button()
            
            # Handle events
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    exit()


            pygame.display.flip()
            clock.tick(self.speed)  # Adjust the tick rate for smoother gameplay  
    
    def Replay_sim(self):
        RES = self.WIDTH, self.HEIGHT
        FONT_SIZE = 34
        


        pygame.init()
        sc = pygame.display.set_mode(RES)
        font = pygame.font.Font(None, FONT_SIZE)
        titlefont = pygame.font.Font(None, 68)
        clock = pygame.time.Clock()

        buttons = [Button(sc, (0, 0, 255), "Next", pygame.Rect(self.WIDTH // 2 - 100, self.HEIGHT // 2 + 200, 200, 50), font, self.replay)]#, Button(sc, (0, 0, 255), "Back", pygame.Rect(self.WIDTH // 2 - 100, self.HEIGHT // 2 + 300, 200, 50), self.replay)
        while True:
            #sc.fill((30, 30, 30))
            sc.fill((0, 0, 0)) #cool mode
            #sc.fill((255, 255, 255)) try it I DARE YOU
            #for x in texts:
                #x.draw_text()
            for x in buttons:
                x.draw_button()
            
            # Handle events
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    exit()


            pygame.display.flip()
            clock.tick(self.speed)  # Adjust the tick rate for smoother gameplay  
    #this was a worth it change